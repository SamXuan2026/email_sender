#!/usr/bin/env python3
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from flask_cors import CORS
import csv
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
import datetime
import json

app = Flask(__name__, static_folder='../dist', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)
HOST_IP = '172.16.1.32'

# 启用CORS支持
CORS(app, resources={r"/*": {"origins": "*"}})

# 保存上传文件的目录
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 保存临时数据的目录
DATA_FOLDER = 'data'
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)

# 数据文件路径
DATA_FILE = os.path.join(DATA_FOLDER, 'current_data.json')

# 全局变量，用于存储当前上传的数据
current_data = {
    'rows': [],
    'email_column': 'email',
    'filename': '',
    'uploaded_file': ''
}

try:
    from openpyxl import load_workbook
except ImportError:
    print("Warning: openpyxl not installed. Excel files will not be supported.")
    load_workbook = None

# 加载保存的数据
def load_data():
    global current_data
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                import json
                current_data = json.load(f)
            print('数据加载成功')
        except Exception as e:
            print(f'数据加载失败: {str(e)}')

# 保存数据
def save_data():
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            import json
            json.dump(current_data, f, ensure_ascii=False, indent=2)
        print('数据保存成功')
    except Exception as e:
        print(f'数据保存失败: {str(e)}')

# 初始化时加载数据
load_data()

def send_email(sender_email, sender_password, recipient_email, subject, body, smtp_server, smtp_port):
    """发送纯文本邮件"""
    try:
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 根据端口号决定使用哪种连接方式
        if smtp_port == 465:  # SSL端口
            with smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30) as server:
                server.login(sender_email, sender_password)
                server.send_message(msg)
        else:  # STARTTLS端口
            with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)


def send_html_email(sender_email, sender_password, recipient_email, subject, html_body, smtp_server, smtp_port):
    """发送HTML邮件"""
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = f"人事小助手 <{sender_email}>"
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # 添加HTML内容
        html_part = MIMEText(html_body, 'html', 'utf-8')
        msg.attach(html_part)
        
        # 根据端口号决定使用哪种连接方式
        if smtp_port == 465:  # SSL端口
            with smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30) as server:
                server.login(sender_email, sender_password)
                server.send_message(msg)
        else:  # STARTTLS端口
            with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)

def read_excel_file(file_path, email_column):
    """读取Excel文件"""
    rows = []
    try:
        if load_workbook is None:
            return rows, "openpyxl not installed. Cannot read Excel files."
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 找到真正的表头行（不是全为None的行）
        headers = None
        header_row = 1
        
        for i in range(1, min(10, ws.max_row + 1)):  # 最多检查前10行
            row_values = [cell.value for cell in ws[i]]
            # 过滤掉None值
            non_none_values = [v for v in row_values if v is not None]
            if len(non_none_values) > 0:
                headers = row_values
                header_row = i
                break
        
        if headers is None:
            return rows, "没有找到表头行"
        
        # 从表头行的下一行开始读取数据
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # 过滤掉空行
            non_none_values = [v for v in row if v is not None]
            if len(non_none_values) > 0:
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                rows.append(row_dict)
    except Exception as e:
        return rows, str(e)
    return rows, ""

@app.route('/')
def index():
    return send_from_directory('../dist', 'index.html')

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        # 保存设置到会话或数据库
        smtp_server = request.form.get('smtp_server')
        smtp_port = request.form.get('smtp_port')
        sender_email = request.form.get('sender_email')
        # sender_password 仅用于当前请求，不在持久存储中保存
        sender_password = request.form.get('sender_password')
        
        # 这里可以将设置保存到数据库或配置文件
        # 暂时使用会话存储，但不保存敏感密码
        session['smtp_server'] = smtp_server
        session['smtp_port'] = smtp_port
        session['sender_email'] = sender_email
        
        flash('设置保存成功！')
        return redirect(url_for('settings'))
    
    # 获取默认值
    smtp_server = session.get('smtp_server', 'smtp.gmail.com')
    smtp_port = session.get('smtp_port', '587')
    sender_email = session.get('sender_email', '')
    sender_password = ''
    
    return send_from_directory('../dist', 'index.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files.get('file')
        email_column = request.form.get('email_column', 'email')
        
        if not file:
            flash('请选择文件！')
            return redirect(url_for('upload'))
        
        # 保存上传的文件
        filename = file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # 读取文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            rows, error = read_excel_file(file_path, email_column)
        else:
            flash('不支持的文件格式，请使用Excel文件！')
            return redirect(url_for('upload'))
        
        if error:
            flash(f'读取文件失败: {error}')
            return redirect(url_for('upload'))
        
        if not rows:
            flash('没有读取到数据！')
            return redirect(url_for('upload'))
        
        # 将rows转换为可序列化的格式，处理非JSON可序列化的数据类型
        def serialize_value(value):
            if value is None:
                return None
            elif isinstance(value, (datetime.datetime, datetime.date)):
                return value.isoformat()
            elif isinstance(value, datetime.time):
                return value.strftime('%H:%M:%S')
            else:
                return str(value)
        
        # 处理每一行数据
        serialized_rows = []
        for row in rows:
            serialized_row = {}
            for key, value in row.items():
                serialized_row[key] = serialize_value(value)
            serialized_rows.append(serialized_row)
        
        # 限制数据大小
        max_rows = 100  # 最多保存100行数据
        if len(serialized_rows) > max_rows:
            current_data['rows'] = serialized_rows[:max_rows]
            print(f'数据行数超过限制，只保存前{max_rows}行')
        else:
            current_data['rows'] = serialized_rows
        
        current_data['email_column'] = email_column
        current_data['filename'] = filename
        current_data['uploaded_file'] = file_path
        
        # 保存数据到文件
        save_data()
        
        return redirect(url_for('preview'))
    
    return send_from_directory('../dist', 'index.html')

@app.route('/preview')
def preview():
    rows = current_data['rows']
    email_column = current_data['email_column']
    filename = current_data['filename']
    
    if not rows:
        flash('没有数据可预览！')
        return redirect(url_for('upload'))
    
    # 获取列名
    if rows:
        columns = list(rows[0].keys())
    else:
        columns = []
    
    return send_from_directory('../dist', 'index.html')

@app.route('/send_single', methods=['POST'])
def send_single():
    # 获取表单数据
    row_index = request.form.get('row_index', type=int)
    subject = request.form.get('subject', '您的工资单')
    
    # 获取SMTP设置
    smtp_server = request.form.get('smtp_server', 'smtp.gmail.com')
    smtp_port = int(request.form.get('smtp_port', '587'))
    sender_email = request.form.get('sender_email')
    sender_password = request.form.get('sender_password')
    
    # 获取数据
    rows = current_data['rows']
    email_column = current_data['email_column']
    
    if not rows or row_index >= len(rows):
        return jsonify({'success': False, 'message': '数据不存在！'})
    
    if not sender_email or not sender_password:
        return jsonify({'success': False, 'message': '请填写发件人邮箱和密码！'})
    
    # 获取当前行数据
    row = rows[row_index]
    email = row.get(email_column)
    
    if not email:
        return jsonify({'success': False, 'message': '该行数据没有邮箱地址！'})
    
    # 构建HTML邮件内容
    table_rows = ""
    for key, value in row.items():
        if key != email_column:
            table_rows += f'<tr><td>{key}</td><td>{value}</td><td></td></tr>'
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
            .email-container {{ max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            .header {{ text-align: center; padding: 20px 0; border-bottom: 2px solid #4CAF50; }}
            .header h1 {{ color: #333; margin: 0; font-size: 24px; }}
            .greeting {{ font-size: 16px; color: #555; margin: 20px 0; }}
            .data-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .data-table th, .data-table td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            .data-table th {{ background-color: #4CAF50; color: white; }}
            .data-table tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .signature {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; color: #777; font-style: italic; }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <div class="header">
                <h1>{subject}</h1>
            </div>
            <div class="greeting">
                <p>尊敬的用户，您好：</p>
                <p>以下是您的数据信息：</p>
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>字段</th>
                        <th>值</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <div class="signature">
                <p>此致</p>
                <p>人事小助手</p>
                <p>如有疑问，请回复此邮件。</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # 发送邮件
    print(f"准备发送邮件到: {email}")
    print(f"SMTP服务器: {smtp_server}:{smtp_port}")
    print(f"发件人邮箱: {sender_email}")
    
    success, error = send_html_email(
        sender_email=sender_email,
        sender_password=sender_password,
        recipient_email=email,
        subject=subject,
        html_body=html_body,
        smtp_server=smtp_server,
        smtp_port=smtp_port
    )
    
    if success:
        print(f"邮件发送成功到: {email}")
        return jsonify({'success': True, 'message': f'邮件发送成功到: {email}'})
    else:
        print(f"邮件发送失败到: {email}. 错误: {error}")
        return jsonify({'success': False, 'message': f'邮件发送失败到: {email}. 错误: {error}'})

@app.route('/upload_data', methods=['POST'])
def upload_data():
    """接收前端上传的数据"""
    global current_data
    
    # 获取数据
    email_column = request.form.get('email_column', 'email')
    data_json = request.form.get('data')
    
    if not data_json:
        return jsonify({'success': False, 'message': '没有收到数据！'})
    
    try:
        rows = json.loads(data_json)
        
        # 限制数据大小
        max_rows = 100  # 最多保存100行数据
        if len(rows) > max_rows:
            current_data['rows'] = rows[:max_rows]
            print(f'数据行数超过限制，只保存前{max_rows}行')
        else:
            current_data['rows'] = rows
        
        current_data['email_column'] = email_column
        current_data['filename'] = 'uploaded_data'
        current_data['uploaded_file'] = 'uploaded_data'
        
        # 保存数据到文件
        save_data()
        
        print(f'数据上传成功，共{len(rows)}行，邮箱列: {email_column}')
        return jsonify({'success': True, 'message': '数据上传成功！'})
    except Exception as e:
        print(f'数据处理失败: {str(e)}')
        return jsonify({'success': False, 'message': f'数据处理失败: {str(e)}'})

@app.route('/send_test_email', methods=['POST'])
def send_test_email():
    """发送测试邮件"""
    # 获取表单数据
    recipient_email = request.form.get('recipient_email', 'xuanjueming@foreverht.com')
    subject = request.form.get('subject', '您的工资单')
    
    # 获取SMTP设置
    smtp_server = request.form.get('smtp_server', 'smtp.gmail.com')
    smtp_port = int(request.form.get('smtp_port', '587'))
    sender_email = request.form.get('sender_email')
    sender_password = request.form.get('sender_password')
    
    if not sender_email or not sender_password:
        return jsonify({'success': False, 'message': '请填写发件人邮箱和密码！'})
    
    # 构建HTML邮件内容
    html_body = """
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body { font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }
            .email-container { max-width: 600px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            .header { text-align: center; padding: 20px 0; border-bottom: 2px solid #4CAF50; }
            .header h1 { color: #333; margin: 0; font-size: 24px; }
            .greeting { font-size: 16px; color: #555; margin: 20px 0; }
            .salary-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
            .salary-table th, .salary-table td { border: 1px solid #ddd; padding: 12px; text-align: left; }
            .salary-table th { background-color: #4CAF50; color: white; }
            .salary-table tr:nth-child(even) { background-color: #f2f2f2; }
            .signature { margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; color: #777; font-style: italic; }
        </style>
    </head>
    <body>
        <div class="email-container">
            <div class="header">
                <h1>您的工资单</h1>
            </div>
            <div class="greeting">
                <p>尊敬的员工，您好：</p>
                <p>这是您的工资单详情，请查收。</p>
            </div>
            <table class="salary-table">
                <thead>
                    <tr>
                        <th>项目</th>
                        <th>金额</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>基本工资</td>
                        <td>15,000.00</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>绩效奖金</td>
                        <td>3,000.00</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>餐补</td>
                        <td>500.00</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>交通补助</td>
                        <td>800.00</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>社保个人部分</td>
                        <td>-1,200.00</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td>公积金个人部分</td>
                        <td>-1,500.00</td>
                        <td></td>
                    </tr>
                    <tr style="background-color: #e8f5e8; font-weight: bold;">
                        <td>实发工资</td>
                        <td>16,600.00</td>
                        <td></td>
                    </tr>
                </tbody>
            </table>
            <div class="signature">
                <p>此致</p>
                <p>人事小助手</p>
                <p>如有疑问，请联系人事部门</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # 发送邮件
    print(f"准备发送测试邮件到: {recipient_email}")
    print(f"SMTP服务器: {smtp_server}:{smtp_port}")
    print(f"发件人邮箱: {sender_email}")
    
    success, error = send_html_email(
        sender_email=sender_email,
        sender_password=sender_password,
        recipient_email=recipient_email,
        subject=subject,
        html_body=html_body,
        smtp_server=smtp_server,
        smtp_port=smtp_port
    )
    
    if success:
        print(f"测试邮件发送成功到: {recipient_email}")
        return jsonify({'success': True, 'message': f'测试邮件发送成功到: {recipient_email}'})
    else:
        print(f"测试邮件发送失败到: {recipient_email}. 错误: {error}")
        return jsonify({'success': False, 'message': f'测试邮件发送失败到: {recipient_email}. 错误: {error}'})

@app.route('/clear_session')
def clear_session():
    """清除会话数据"""
    # 清除会话数据
    session.clear()
    # 清除全局数据
    global current_data
    current_data = {
        'rows': [],
        'email_column': 'email',
        'filename': '',
        'uploaded_file': ''
    }
    # 保存空数据到文件
    save_data()
    # 显示成功消息
    flash('历史会话已清除！', 'success')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host=HOST_IP, port=5004)
